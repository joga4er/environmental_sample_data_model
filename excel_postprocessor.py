from linkml_runtime.utils.schemaview import SchemaView
from typing import Any

import openpyxl as xlprocessor
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.cell import get_column_letter

def standard_code_validation_type(cell: str) -> DataValidation:
    """Creates Validation formula for standard code for certain excel cell."""
    custom_standard_code_validation = DataValidation(
        type="custom", formula1=f'=EXACT(LEFT({cell},1),"#")', showErrorMessage=True,
    )

    # Set the error message
    custom_standard_code_validation.error = "Please enter a value starting with #."
    custom_standard_code_validation.errorTitle = "Input Error"
    custom_standard_code_validation.prompt = "Please enter a value starting with #."
    custom_standard_code_validation.promptTitle = "Input Requirement"

    return custom_standard_code_validation

def add_compulsory_fields_validation_colum(sheet: xlprocessor.Workbook, compulsory_fields: list[int], colidx: int) -> None:
    """Adds colum which checks if all compulsory fields are present."""
    sheet.cell(1, colidx + 1).value = 'Required fields'
    sheet.cell(2, colidx + 1).value = 'Shows ✓, if all required fields are provided.'

    # write validation formula to each row of validation column
    for rowidx in range(3, 999):
        # create formula for all required columns for each row
        cell_formula = ''
        for (idx, compulsory_field) in enumerate(compulsory_fields):
            # check if len of compulsory field is greater than zero
            this_cell_formula = f'LEN({get_column_letter(compulsory_field)}{rowidx})>0'
            # combine formula for all fields with AND
            if idx == 0:
                cell_formula = this_cell_formula
            else:
                cell_formula = f'AND({cell_formula},{this_cell_formula})'
        # write formulat to last column of row
        sheet.cell(rowidx, colidx + 1).value = f'=IF({cell_formula},"✓","x")'

def add_unique_fields_validation_column(sheet:xlprocessor.Workbook, relevant_column: int, relevant_column_range:list[str], colidx:int) -> None:
    """Adds colum which checks if all field values are unique and occur in reference column of other sheet."""
    relevant_column_letter = get_column_letter(relevant_column)
    sheet.cell(1, colidx + 1).value = f'Reference {relevant_column_letter}'
    sheet.cell(2, colidx + 1).value = f'Shows ✓, if all required fields of column {relevant_column_letter} are unique and provided in reference column: {relevant_column_range}.'

    for rowidx in range(3, 999):
        cell_formula = ''
        for (idx, column_range) in enumerate(relevant_column_range):
            # check if len of compulsory field is greater than zero
            this_cell_formula = f'NOT(ISERROR(MATCH({relevant_column_letter}{rowidx},{column_range},0)))'
            # combine formula for all fields with AND
            if idx == 0:
                cell_formula = this_cell_formula
            else:
                cell_formula = f'OR({this_cell_formula},{cell_formula})'
        sheet.cell(rowidx, colidx + 1).value = f'=IF({cell_formula},"✓","x")'

class WorkbookStyles:
    """Defines custom cell formats and has functions to set them according to patterns and ranges."""

    custom_date = NamedStyle(name='custom_date', number_format='YYYY-MM-DD')
    custom_time = NamedStyle(name='custom_time', number_format='HH:MM')
    custom_lon_lat = NamedStyle(name='custom_six_digits', number_format='0.000000')
    custom_lcms_code = NamedStyle(name='custom_lcms_code', number_format='00\/0000')
    custom_string = NamedStyle(name='custom_string', number_format='@')
    custom_integer = NamedStyle(name='custom_integer', number_format='0')

    def __init__(self, sheet: xlprocessor):
        """initialize class instance with sheet."""
        self.sheet = sheet
        
    def apply_patterns_and_ranges(self, field_attributes: dict, colidx: int) -> None:
        """Translates patterns and ranges to excel - highly customized, function is not easy to generalize."""
        # set cell format according to field ranges
        if field_attributes.range is None:        
            # apply rules for dates, latitude, longitude, lcms code and standard codes.
            if field_attributes.pattern == '[/d{4}-/d{2}-/d{2}]':
                for rowidx in range(3, 999):
                    self.sheet.cell(rowidx, colidx + 1).style = self.custom_date
            elif field_attributes.pattern == '[/d{4}:/d{2}]':
                for rowidx in range(3, 999):
                    self.sheet.cell(rowidx, colidx + 1).style = self.custom_time
            elif field_attributes.pattern in ['[/d{1,2}./d{6}]', '[/d{1,2,3}./d{6}]']:
                for rowidx in range(3, 999):
                    self.sheet.cell(rowidx, colidx + 1).style = self.custom_lon_lat
            elif field_attributes.pattern == '[/d{2}///d{4}]':
                for rowidx in range(3, 999):
                    self.sheet.cell(rowidx, colidx + 1).style = self.custom_lcms_code
            elif field_attributes.pattern == '[/#+$]':
                cell_column_letter = get_column_letter(colidx + 1)
                for rowidx in range(3, 999):
                    cell = cell_column_letter + str(rowidx)
                    custom_standard_code_validation = standard_code_validation_type(cell=cell)
                    self.sheet.add_data_validation(custom_standard_code_validation)
                    custom_standard_code_validation.add(self.sheet.cell(rowidx, colidx + 1))
            else:
                for rowidx in range(3, 999):
                    self.sheet.cell(rowidx, colidx + 1).style = self.custom_string
        elif field_attributes.range == 'integer':
            for rowidx in range(3, 999):
                self.sheet.cell(rowidx, colidx + 1).style = self.custom_integer
        else:
            pass

def excel_postprocessing(filename: str, slots: Any) -> None:  
    """Iterates over excel schema columns, adds field name descriptions, and marks required fields.
    Adds validation column to check if all mandatory fields have been provided.
    Changes cell formats and adds validation rules to ensure ranges and patterns match linkml schema.
    Adds validation columns to check that linked identifiers are available for each location and each collection."""

    # open excel
    wb = xlprocessor.load_workbook(filename)
    # iterate over sheets
    for sheet in wb:
        compulsory_fields = []
        wb_styler = WorkbookStyles(sheet=sheet)
        # Access and manipulate data in the sheet -> iterate over columns in each sheet
        for (colidx, col) in enumerate(sheet[sheet.min_row : sheet.max_row]):
            field_attributes = slots[col.value]
            # mark required fields
            if field_attributes.required:
                col.value = col.value + ' *'
                col.font = xlprocessor.styles.Font(bold=True)
                compulsory_fields.append(colidx + 1)
            # add field description
            sheet.cell(2, colidx + 1).value = field_attributes.description
            wb_styler.apply_patterns_and_ranges(field_attributes=field_attributes, colidx=colidx)
 
        if colidx > 0:
            add_compulsory_fields_validation_colum(sheet=sheet, colidx=colidx, compulsory_fields=compulsory_fields)
            colidx += 1
        if sheet.title == 'EnvironmentalSample':
            add_unique_fields_validation_column(sheet=sheet, relevant_column_range=['SamplingLocation!A3:A999'],relevant_column=8,colidx=colidx)
            add_unique_fields_validation_column(sheet=sheet, relevant_column_range=[
                'SampleCollection!B3:B999', 'WaterSampleCollection!E3:E999', 'BiotaSampleCollection!G3:G999',
            ], relevant_column=9,colidx=colidx + 1)
    wb.save(filename=filename)
    wb.close()
